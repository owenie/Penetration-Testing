import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from lxml import etree
import pymysql
import pymysql.cursors


try:
    import cookielib
except:
    import http.cookiejar as cookielib
import os.path
import json

baseurl = "https://movie.douban.com/top250/"


headers = {"Host": "www.douban.com",
           "Referer": "https://www.douban.com/",
           'User-Agent': 'Mozilla/5.0 (Windows NT 20.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.226 Safari/537.36 Edge/25.25063'
           }
headers2 = {
    "Host": "movie.douban.com",
    "Referer": "https://www.douban.com/",
    'User-Agent': 'Mozilla/5.0 (Windows NT 20.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.226 Safari/537.36 Edge/25.25063',
    'Connection': 'keep-alive',
}
# 使用cookie登录信息
session = requests.session()  # session
session.cookies = cookielib.LWPCookieJar(filename='cookies')

try:
    session.cookies.load(ignore_discard=True)
    print('成功加载cookie!')
except:
    print("cookie 未能加载!")


def get_captcha(url):
    # 获取验证码
    # pillow 的 Image 显示验证码
    print('成功获取验证码', url)
    captcha_url = url
    response = session.get(captcha_url, headers=headers)
    # 保存验证码
    with open('captcha.jpg', 'wb') as f:
        f.write(response.content)
        f.close()
    try:
        image = Image.open('captcha.jpg')
        image.show()  # 主动打开验证码图片
        image.close()
    except:
        print('请到 %s 目录找到captcha.jpg 手动输入' % os.path.abspath('captcha.jpg'))
    captcha = input("请输入验证码\n>")
    return captcha


def login(acount, secret):
    homeurl = "https://www.douban.com/"
    htmlcha = session.get(homeurl, headers=headers).text
    # 带验证码请求
    patterncha = r'id="captcha_image" src="(.*?)" alt="captcha"'
    httpcha = re.findall(patterncha, htmlcha)
    pattern2 = r'type="hidden" name="captcha-id" value="(.*?)"'
    hidden_value = re.findall(pattern2, htmlcha)
    print(hidden_value)

    data = {
        "source": "index_nav",
        'form_email': acount,
        'form_password': secret
    }
    # 模拟登录请求
    if len(httpcha) > 0:
        print('验证码连接', httpcha)
        capcha = get_captcha(httpcha[0])
        data['captcha-solution'] = capcha
        data['captcha-id'] = hidden_value[0]

    print(data)
    post_url = 'https://www.douban.com/accounts/login'
    login_page = session.post(post_url, data=data, headers=headers)
    # 保存cookies
    session.cookies.save()
    print('登录成功，开始爬取数据。')

    # if isLogin():
    #     print('登录成功，开始爬取数据。')
    # else:
    #     print('登录失败！')


def Getbydom(url):
    """获取页面内容"""
    try:
        data = session.get(url, headers=headers2).text
        soup = BeautifulSoup(data, 'html.parser')
        ol = soup.find('ol', class_='grid_view')  # 有序列表：<ol class="grid_view">
        name = []  # 名称
        star_con = []  # 评价人数
        score = []  # 评分
        info = []  # 信息
        comment_list = []  # 短评
        for i in ol.find_all('li'):
            detail = i.find('div', attrs={'class': 'hd'})
            name_item = detail.find(
                'span', attrs={'class': 'title'}).get_text()  # 电影名称
            score_item = i.find(
                'span', attrs={'class': 'rating_num'}).get_text()  # 评分

            star_item = i.find('div', attrs={'class': 'star'}).find(
                text=re.compile('评价'))  # 评价人数
            info_item = i.find('div', attrs={'class': 'bd'}
                               ).find('p').get_text().strip()  # 电影信息

            info_item = "".join(info_item.split())
            comment = i.find('span', attrs={'class': 'inq'})  # 评论
            # print(info)
            # 添加
            if comment:  # 判断
                comment_list.append(comment.get_text())
            else:
                comment_list.append('无')
            # 单页数据缓存列表
            score.append(score_item)
            name.append(name_item)
            star_con.append(star_item)
            info.append(info_item)

    except Exception as e:
        print(e)

    return name, star_con, score, comment_list, info


def Getbylxml(url):
    """获取页面内容"""
    data = session.get(url, headers=headers2).text
    name = []  # 名称
    star_con = []  # 评价人数
    score = []  # 评分
    info = []  # 信息
    comment_list = []  # 短评
    # s = requests.session()
    # s.keep_alive = False
    response = etree.HTML(data)
    result = response.xpath(
        ".//*[@id='content']/div/div[1]/ol/li")
    for item in result:
        name_item = item.xpath(
            "./div/div[2]/div[1]/a/span[1]/text()")[0].strip()
        info_item = item.xpath("./div/div[2]/div[2]/p[1]/text()")[0].strip()
        score_item = item.xpath(
            "./div/div[2]/div[2]/div/span[2]/text()")[0].strip()
        star_item = item.xpath(
            "./div/div[2]/div[2]/div/span[4]/text()")[0].strip()
        try:
            comment = item.xpath(
                "./div/div[2]/div[2]/p[2]/span/text()")[0].strip()
            comment_list.append(comment)
        except Exception as e:
            comment_list.append('无')

        score.append(score_item)
        name.append(name_item)
        star_con.append(star_item)
        info.append(info_item)
    return name, star_con, score, comment_list, info


def Getall():
    name = []
    star_con = []
    score = []  # 下一个url，爬分页数据
    comment = []
    info = []
    for item in range(0, 10):  # 下一个url，爬分页数据
        nexturl = baseurl + "?start=%d" % (item * 25)
        print("获取URL:" + nexturl)
        # 调用Get函数，爬取数据
        # name_item, star_item, score_item, comment_item, info_item = Getbydom(
        #     nexturl)
        name_item, star_item, score_item, comment_item, info_item = Getbylxml(
            nexturl)
        # 打印
        # 分页数据缓存
        name = name + name_item
        star_con = star_con + star_item
        score = score + score_item
        comment = comment + comment_item
        info = info + info_item
    return name, star_con, score, comment, info


def savexslx(name, star_con, score, comment, info):  # 数据保存到xslx文件
    wb = Workbook()
    filename = '豆瓣电影top250.xlsx'
    ws = wb.active
    ws.title = "豆瓣电影top250"

    ws['A1'] = '电影名称'
    ws['B1'] = '评分人数'
    ws['C1'] = '评分'
    ws['D1'] = '评价'
    ws['E1'] = '电影信息'
    # """保存到xlsx中 """
    for (i, m, o, p, q) in zip(name, star_con, score, comment, info):
        col_A = 'A%s' % (name.index(i) + 2)
        col_B = 'B%s' % (name.index(i) + 2)
        col_C = 'C%s' % (name.index(i) + 2)
        col_D = 'D%s' % (name.index(i) + 2)
        col_E = 'E%s' % (name.index(i) + 2)
        ws[col_A] = i
        ws[col_B] = m
        ws[col_C] = o
        ws[col_D] = p
        ws[col_E] = q

    wb.save(filename=filename)
    print("保存%s成功！" % filename)


def savemysql():  # 数据保存到mysql中
    name, star_con, score, comment, info = Getall()
    db = pymysql.connect(host='localhost', user='root', passwd='root', db='movies',
                         port=3306, charset='utf8', cursorclass=pymysql.cursors.DictCursor)  # 连接数据库
    db.autocommit(True)
    cursor = db.cursor()
    for (i, m, o, p, q) in zip(name, star_con, score, comment, info):
        cursor.execute(
            'insert into moviestop(name, star_con, score, comment, info) values(%s,%s,%s,%s,%s)', [i, m, o, p, q])

    db.close()
    cursor.close()
    print('Save to mysql success！')


if __name__ == '__main__':

    login('账号', '密码')
    name, star_con, score, comment, info = Getall()
    savexslx(name, star_con, score, comment, info)
