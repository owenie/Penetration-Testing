import requests
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from lxml import etree
import pymysql
import pymysql.cursors

baseurl = "http://movie.douban.com/top250/"

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 20.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/52.0.2743.226 Safari/537.36 Edge/25.25063'
}


def Get(url):
    """获取页面内容"""
    try:
        data = requests.get(url, headers=headers).text
        soup = BeautifulSoup(data, 'html.parser')
        ol = soup.find('ol', class_='grid_view')  # 有序列表：<ol class="grid_view">
        name = []  # 名称
        star_con = []  # 评价人数
        score = []  # 评分
        info = []  # 信息
        comment_list = []  # 短评
        for i in ol.find_all('li'):
            detail = i.find('div', attrs={'class': 'hd'})
            name_item = detail.find(
                'span', attrs={'class': 'title'}).get_text()  # 电影名称
            score_item = i.find(
                'span', attrs={'class': 'rating_num'}).get_text()  # 评分

            star_item = i.find('div', attrs={'class': 'star'}).find(
                text=re.compile('评价'))  # 评价人数
            print(star_item)
            info_item = i.find('div', attrs={'class': 'bd'}
                               ).find('p').get_text().strip()  # 电影信息

            info_item = "".join(info_item.split())
            comment = i.find('span', attrs={'class': 'inq'})  # 评论
            # print(info)
            # 添加
            if comment:  # 判断
                comment_list.append(comment.get_text())
            else:
                comment_list.append('无')
            # 单页数据缓存列表
            score.append(score_item)
            name.append(name_item)
            star_con.append(star_item)
            info.append(info_item)

    except Exception as e:
        print(e)

    return name, star_con, score, comment_list, info


def Getbylxml(url):
    """获取页面内容"""
    data = requests.get(url, headers=headers).text
    name = []  # 名称
    star_con = []  # 评价人数
    score = []  # 评分
    info = []  # 信息
    comment_list = []  # 短评
    # s = requests.session()
    # s.keep_alive = False
    response = etree.HTML(data)
    result = response.xpath(
        ".//*[@id='content']/div/div[1]/ol/li")
    for item in result:

        name_item = item.xpath(
            "./div/div[2]/div[1]/a/span[1]/text()")[0].strip()
        info_item = item.xpath("./div/div[2]/div[2]/p[1]/text()")[0].strip()
        score_item = item.xpath(
            "./div/div[2]/div[2]/div/span[2]/text()")[0].strip()
        star_item = item.xpath(
            "./div/div[2]/div[2]/div/span[4]/text()")[0].strip()
        try:
            comment = item.xpath(
                "./div/div[2]/div[2]/p[2]/span/text()")[0].strip()
            comment_list.append(comment)
        except Exception as e:
            comment_list.append('无')

        score.append(score_item)
        name.append(name_item)
        star_con.append(star_item)
        info.append(info_item)

    return name, star_con, score, comment_list, info


def Getall():
    name = []
    star_con = []
    score = []  # 下一个url，爬分页数据
    comment = []
    info = []
    for item in range(0, 10):
         # 下一个url，爬分页数据
        nexturl = baseurl + "?start=%d" % (item * 25)
        print("获取URL:" + nexturl)
        # 调用Get函数，爬取数据
        name_item, star_item, score_item, comment_item, info_item = Getbylxml(
            nexturl)
        # 打印
        # print(name_item, star_item, score_item, comment_item,info_item)
        # 分页数据缓存
        name = name + name_item
        star_con = star_con + star_item
        score = score + score_item
        comment = comment + comment_item
        info = info + info_item
    return name, star_con, score, comment, info


def savexslx(name, star_con, score, comment, info):  # 数据保存到xslx文件
    wb = Workbook()
    filename = '豆瓣电影top250.xlsx'
    ws = wb.active
    ws.title = "豆瓣电影top250"

    ws['A1'] = '电影名称'
    ws['B1'] = '评分人数'
    ws['C1'] = '评分'
    ws['D1'] = '评价'
    ws['E1'] = '电影信息'
    # """保存到xlsx中 """
    for (i, m, o, p, q) in zip(name, star_con, score, comment, info):
        col_A = 'A%s' % (name.index(i) + 2)
        col_B = 'B%s' % (name.index(i) + 2)
        col_C = 'C%s' % (name.index(i) + 2)
        col_D = 'D%s' % (name.index(i) + 2)
        col_E = 'E%s' % (name.index(i) + 2)
        ws[col_A] = i
        ws[col_B] = m
        ws[col_C] = o
        ws[col_D] = p
        ws[col_E] = q

    wb.save(filename=filename)
    print("保存%s成功！" % filename)


def savemysql():  # 数据保存到mysql中
    name, star_con, score, comment, info = Getall()
    db = pymysql.connect(host='localhost', user='root', passwd='root', db='movies',
                         port=3306, charset='utf8', cursorclass=pymysql.cursors.DictCursor)  # 连接数据库
    db.autocommit(True)
    cursor = db.cursor()
    for (i, m, o, p, q) in zip(name, star_con, score, comment, info):
        cursor.execute(
            'insert into moviestop(name, star_con, score, comment, info) values(%s,%s,%s,%s,%s)', [i, m, o, p, q])

    db.close()
    cursor.close()
    print('Save to mysql success！')


if __name__ == '__main__':
    # savemysql()
    Get(baseurl)
    # name, star_con, score, comment, info = Getall()
    # savexslx(name, star_con, score, comment, info)
